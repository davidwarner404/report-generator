from flask import Flask, render_template, request, send_file
import pandas as pd  
import io
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        form_id = request.form.get('form_id')  # Get the form_id from the submitted form
        

        # Ensure the correct files are provided based on the form
        if form_id == 'EM01':
            formater='EM01.xlsx'
            file1 = request.files.get('file1')
            file2 = request.files.get('file2')
        elif form_id == 'EM02':
            formater='EM02.xlsx'
            file1 = request.files.get('file3')
            file2 = request.files.get('file4')
        else:
            return redirect(request.url)  # Redirect if no valid form_id is found

        # Validate that files are uploaded
        if not file1 or not file2 or file1.filename == '' or file2.filename == '':
            return redirect(request.url)

        # Process files if both are provided
        if file1 and file2:
            output_buffer = merge_files(file1, file2,formater)  # Assuming `merge_files` is implemented

            # Return the buffer as a downloadable file
            return send_file(
                output_buffer,
                as_attachment=True,
                download_name=f'merged_file_{form_id}.xlsx',  # Name file based on form_id
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    return render_template('index.html')

def merge_files(file1, file2,formater):
    # Read the files into DataFrames
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Create a pivot table from df1
    pivot_table = pd.pivot_table(df1, index=['Segment', 'Sub Division', 'Report List'],
                             values=['Delivered', 'Open', 'Clicks', 'Unsubscibes', 'Complaints'],
                             aggfunc='sum', fill_value=0)

    # Create a BytesIO buffer
    buffer = io.BytesIO()

    # Save the pivot table to the buffer
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        pivot_table.to_excel(writer, sheet_name='PivotTable', index=True)

    # Ensure the buffer is at the beginning
    buffer.seek(0)

    # Load sheet1 for merging
    wb = load_workbook(formater)
    ws = wb.active

    # Convert the sheet1 data to a pandas DataFrame
    sheet1_df = pd.DataFrame(ws.values)
    sheet1_df.columns = sheet1_df.iloc[0]
    sheet1_df = sheet1_df[1:]

    # Load the pivot table data from the buffer
    buffer.seek(0)
    pivot_df = pd.read_excel(buffer, sheet_name='PivotTable')

    # Ensure relevant columns are properly named
    pivot_df = pivot_df[['Report List', 'Delivered', 'Open', 'Clicks', 'Unsubscibes', 'Complaints']]

    # Update sheet1 values with the pivot table values
    for index, row in sheet1_df.iterrows():
        matching_row = pivot_df[pivot_df['Report List'] == row['Report List']]
        if not matching_row.empty:
            ws.cell(row=index + 1, column=sheet1_df.columns.get_loc('Delivered') + 1).value = matching_row['Delivered'].values[0]
            ws.cell(row=index + 1, column=sheet1_df.columns.get_loc('Open') + 1).value = matching_row['Open'].values[0]
            ws.cell(row=index + 1, column=sheet1_df.columns.get_loc('Clicks') + 1).value = matching_row['Clicks'].values[0]
            ws.cell(row=index + 1, column=sheet1_df.columns.get_loc('Unsubscibes') + 1).value = matching_row['Unsubscibes'].values[0]
            ws.cell(row=index + 1, column=sheet1_df.columns.get_loc('Complaints') + 1).value = matching_row['Complaints'].values[0]

    # Process df2 to add List_id
    df2['Source (Sub Publisher)'] = df2['Source (Sub Publisher)'].astype(str)
    df2['List_id'] = df2['Source (Sub Publisher)'].apply(lambda x: int(x.split('_')[-1]) if (x.split('_')[-1]).isdigit() else x.split('_')[-1])

    # Ensure numeric columns are properly handled
    df2['Unique Clicks'] = pd.to_numeric(df2['Unique Clicks'], errors='coerce')
    df2['Gross Clicks'] = pd.to_numeric(df2['Gross Clicks'], errors='coerce')
    df2['Payout'] = pd.to_numeric(df2['Payout'], errors='coerce')

    # Rearrange columns to place List_id next to Source (Sub Publisher)
    columns = list(df2.columns)
    source_index = columns.index('Source (Sub Publisher)')
    columns.insert(source_index + 1, columns.pop(-1))
    df2 = df2[columns]

    # Group df2 by List_id
    df2_grouped = df2.groupby('List_id').agg({'Unique Clicks': 'sum', 'Gross Clicks': 'sum', 'Payout': 'sum'}).reset_index()

    # Convert grouped data to dictionary
    sheet2_dict = df2_grouped.set_index('List_id').to_dict(orient='index')

    # Update sheet1 with values from df2_grouped
    for index, row in sheet1_df.iterrows():
        list_id = row['List_id']
        if list_id in sheet2_dict:
            data = sheet2_dict[list_id]
            ws.cell(row=index + 1, column=sheet1_df.columns.get_loc('Unique Clicks') + 1).value = data['Unique Clicks']
            ws.cell(row=index + 1, column=sheet1_df.columns.get_loc('Gross Clicks') + 1).value = data['Gross Clicks']
            ws.cell(row=index + 1, column=sheet1_df.columns.get_loc('Payout') + 1).value = data['Payout']

    # Save the updated workbook to the buffer
    buffer.seek(0)
    wb.save(buffer)
    buffer.seek(0)

    return buffer

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)